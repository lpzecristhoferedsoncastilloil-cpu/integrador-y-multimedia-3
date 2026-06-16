import sys
import os
import random
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Asegurar conexion con base de datos a traves de Django
from django.db import connection

def generate_custom_exercise(word, read_as, confusion_key):
    juego = "La Caza del Grafema"
    word_lower = word.lower()
    if len(word) > 8:
        juego = "El Almacén"
    elif "r" in word_lower:
        juego = "El Eco de las Sílabas"
    
    options = []
    if confusion_key == 'B_V':
        options = ["ba", "va", "bo", "vo"]
    elif confusion_key == 'G_J':
        options = ["ga", "ja", "ge", "je"]
    elif confusion_key == 'C_S_Z':
        options = ["ca", "sa", "za", "ce"]
    elif confusion_key == 'MIRROR':
        options = ["da", "ba", "pa", "qa"]
    elif confusion_key == 'R_RR':
        options = ["ra", "rra", "ro", "rro"]
    else:
        if len(word_lower) >= 4:
            options = [word_lower[:2], word_lower[2:4], "la", "te"]
        else:
            options = [word_lower, "xo", "za", "mi"]
            
    options = list(dict.fromkeys(options))[:4]
    while len(options) < 4:
        options.append("ma")
        
    return {
        "palabra": word,
        "opciones": options,
        "juego_recomendado": juego
    }

# Mapeos de juegos a rutas lectoras clinicas
PHONOLOGICAL_GAMES = ['rocket_builder', 'cheese', 'temple', 'train', 'silabas_magicas']
SUPERFICIAL_GAMES = ['grafema', 'hangman', 'warehouse', 'machine']
MIXED_GAMES = ['river', 'maze']

# Etiquetas de confusiones clinicas de grafemas
ETIQUETAS = {
    'B_V': 'Confusión B / V',
    'G_J': 'Confusión G / J',
    'C_S_Z': 'Seseo / Ortografía C-S-Z',
    'MIRROR': 'Letras Espejo (d-b-p-q)',
    'H_OMISSION': 'Omisión de la H',
    'R_RR': 'Grafía R / RR',
    'OTHER': 'Otros fallos visuales/fonológicos'
}

# Ejercicios correctivos predefinidos para confusiones
EXERCISE_TEMPLATES = {
    'B_V': [
        {"palabra": "vaca", "opciones": ["va", "ba", "ka", "ca"], "juego_recomendado": "El Almacén"},
        {"palabra": "bola", "opciones": ["vo", "bo", "lla", "la"], "juego_recomendado": "La Máquina de Sílabas"},
        {"palabra": "tubo", "opciones": ["tu", "vu", "bo", "vo"], "juego_recomendado": "La Caza del Grafema"},
        {"palabra": "barco", "opciones": ["var", "bar", "co", "go"], "juego_recomendado": "El Eco de las Sílabas"},
        {"palabra": "viento", "opciones": ["vien", "bien", "to", "do"], "juego_recomendado": "La Caza del Grafema"}
    ],
    'G_J': [
        {"palabra": "gente", "opciones": ["jen", "gen", "te", "de"], "juego_recomendado": "La Caza del Grafema"},
        {"palabra": "jirafa", "opciones": ["gi", "ji", "ra", "fa"], "juego_recomendado": "El Almacén"},
        {"palabra": "ejemplo", "opciones": ["ejem", "egem", "plo", "blo"], "juego_recomendado": "La Máquina de Sílabas"},
        {"palabra": "gemelos", "opciones": ["je", "ge", "me", "los"], "juego_recomendado": "El Rescate de las Letras"},
        {"palabra": "reloj", "opciones": ["re", "loj", "log", "lox"], "juego_recomendado": "El Eco de las Sílabas"}
    ],
    'C_S_Z': [
        {"palabra": "zapato", "opciones": ["sa", "za", "ca", "pa"], "juego_recomendado": "El Almacén"},
        {"palabra": "casa", "opciones": ["ca", "za", "sa", "ma"], "juego_recomendado": "El Reto del Queso"},
        {"palabra": "cielo", "opciones": ["cie", "sie", "zie", "lo"], "juego_recomendado": "La Caza del Grafema"},
        {"palabra": "zorro", "opciones": ["so", "zo", "rro", "ro"], "juego_recomendado": "El Eco de las Sílabas"},
        {"palabra": "sopa", "opciones": ["so", "zo", "co", "pa"], "juego_recomendado": "El Almacén"}
    ],
    'MIRROR': [
        {"palabra": "dado", "opciones": ["da", "ba", "pa", "do"], "juego_recomendado": "El Almacén"},
        {"palabra": "dedo", "opciones": ["de", "be", "pe", "do"], "juego_recomendado": "La Caza del Grafema"},
        {"palabra": "bota", "opciones": ["bo", "do", "po", "ta"], "juego_recomendado": "El Rescate de las Letras"},
        {"palabra": "pelo", "opciones": ["pe", "be", "de", "lo"], "juego_recomendado": "La Máquina de Sílabas"},
        {"palabra": "queso", "opciones": ["que", "de", "be", "so"], "juego_recomendado": "El Almacén"}
    ],
    'H_OMISSION': [
        {"palabra": "huevo", "opciones": ["hue", "ue", "vo", "bo"], "juego_recomendado": "El Rescate de las Letras"},
        {"palabra": "helado", "opciones": ["he", "e", "la", "do"], "juego_recomendado": "La Caza del Grafema"},
        {"palabra": "hilo", "opciones": ["hi", "i", "lo", "to"], "juego_recomendado": "La Máquina de Sílabas"},
        {"palabra": "hora", "opciones": ["ho", "o", "ra", "da"], "juego_recomendado": "El Almacén"},
        {"palabra": "deshacer", "opciones": ["des", "ha", "cer", "acer"], "juego_recomendado": "La Máquina de Sílabas"}
    ],
    'R_RR': [
        {"palabra": "perro", "opciones": ["pe", "rro", "ro", "lo"], "juego_recomendado": "El Reto del Queso"},
        {"palabra": "pero", "opciones": ["pe", "ro", "rro", "do"], "juego_recomendado": "La Caza del Grafema"},
        {"palabra": "carro", "opciones": ["ca", "rro", "ro", "to"], "juego_recomendado": "El Rescate de las Letras"},
        {"palabra": "caro", "opciones": ["ca", "ro", "rro", "go"], "juego_recomendado": "El Eco de las Sílabas"},
        {"palabra": "correr", "opciones": ["co", "rrer", "rer", "der"], "juego_recomendado": "La Máquina de Sílabas"}
    ],
    'OTHER': [
        {"palabra": "plato", "opciones": ["pla", "pal", "to", "do"], "juego_recomendado": "El Eco de las Sílabas"},
        {"palabra": "fruta", "opciones": ["fru", "fur", "ta", "da"], "juego_recomendado": "La Máquina de Sílabas"},
        {"palabra": "globo", "opciones": ["glo", "gol", "bo", "po"], "juego_recomendado": "La Caza del Grafema"},
        {"palabra": "libro", "opciones": ["li", "bro", "bor", "pro"], "juego_recomendado": "El Almacén"},
        {"palabra": "sol", "opciones": ["sol", "los", "col", "tol"], "juego_recomendado": "El Reto del Queso"}
    ]
}

def analyze_confusion(word_shown, answer_given, error_type=None):
    """
    Analiza la palabra mostrada y la respuesta dada para clasificar el tipo de confusion de grafemas.
    """
    if error_type:
        if 'b_v' in error_type.lower():
            return 'B_V'
        if 'g_j' in error_type.lower():
            return 'G_J'
        if 'c_s_z' in error_type.lower():
            return 'C_S_Z'
        if 'espejo' in error_type.lower() or 'mirror' in error_type.lower():
            return 'MIRROR'
        if 'h' in error_type.lower() and 'omision' in error_type.lower():
            return 'H_OMISSION'
        if 'r_rr' in error_type.lower() or 'r' in error_type.lower():
            return 'R_RR'
            
    w = str(word_shown).upper().replace('_', '')
    a = str(answer_given).upper().replace('_', '')
    
    # 1. B vs V
    if ('B' in w and 'V' in a) or ('V' in w and 'B' in a):
        return 'B_V'
    # 2. G vs J
    if ('G' in w and 'J' in a) or ('J' in w and 'G' in a):
        return 'G_J'
    # 3. C vs S vs Z
    if ('C' in w or 'S' in w or 'Z' in w) and ('C' in a or 'S' in a or 'Z' in a) and w != a:
        return 'C_S_Z'
    # 4. Mirror letters (B, D, P, Q)
    mirror_set = {'B', 'D', 'P', 'Q'}
    w_mirrors = mirror_set.intersection(set(w))
    a_mirrors = mirror_set.intersection(set(a))
    if w_mirrors and a_mirrors and w_mirrors != a_mirrors:
        return 'MIRROR'
    # 5. H omission
    if 'H' in w and 'H' not in a:
        return 'H_OMISSION'
    # 6. R vs RR
    if 'R' in w and 'R' in a and w.count('R') != a.count('R'):
        return 'R_RR'
        
    return 'OTHER'


def get_patient_profile(patient_id):
    """
    Extrae toda la informacion de intentos y resultados del paciente desde la base de datos,
    incluye las evaluaciones de test de dislexia y ejecuta las inferencias del sistema experto.
    """
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id_paciente, nombre_completo, edad_actual, colegio_ocupacion, observaciones_generales, fecha_registro
            FROM pacientes WHERE id_paciente = %s
        """, [patient_id])
        pac_row = cursor.fetchone()
        if not pac_row:
            return None
            
        paciente_info = {
            "id": pac_row[0],
            "nombre": pac_row[1],
            "edad": pac_row[2],
            "colegio": pac_row[3] or "No registrado",
            "observaciones": pac_row[4] or "Sin observaciones",
            "fecha_registro": pac_row[5].strftime("%d/%m/%Y") if pac_row[5] else "N/A"
        }
        
        # Obtener los resultados del test de dislexia para este paciente
        cursor.execute("""
            SELECT id, metodo, a_p, t_p, il_p, r_p, a_ps, t_ps, il_ps, r_ps,
                   diagnostico, detalles_errores, observaciones, fecha_registro
            FROM test_dislexia_results
            WHERE id_paciente = %s
            ORDER BY fecha_registro DESC
        """, [patient_id])
        dislexia_test_rows = cursor.fetchall()
        
        # Obtener los IDs de game_players correspondientes al paciente
        cursor.execute("SELECT id, nickname FROM game_players WHERE id_paciente = %s", [patient_id])
        player_rows = cursor.fetchall()
        player_ids = [row[0] for row in player_rows]
        nicknames = [row[1] for row in player_rows]
        
    latest_test = None
    dis_test_errors = []
    
    if dislexia_test_rows:
        row = dislexia_test_rows[0]
        latest_test = {
            "id": row[0],
            "metodo": row[1],
            "a_p": row[2],
            "t_p": row[3],
            "il_p": row[4],
            "r_p": row[5],
            "a_ps": row[6],
            "t_ps": row[7],
            "il_ps": row[8],
            "r_ps": row[9],
            "diagnostico": row[10],
            "fecha": row[13].strftime("%d/%m/%Y") if row[13] else "N/A"
        }
        
        # Parsear detalles_errores para obtener stutters/vacilaciones o fallos ortográficos
        try:
            if row[11]:
                detalles_list = json.loads(row[11])
                for item in detalles_list:
                    if item.get("state") in ["incorrect", "hesitation"]:
                        dis_test_errors.append(item)
        except Exception:
            pass

    if not player_ids:
        # Fallback si no hay juegos pero sí test
        fon_prob = 0.0
        sup_prob = 0.0
        mix_prob = 0.0
        test_diag = latest_test["diagnostico"] if latest_test else None
        if test_diag:
            if test_diag == "DISLEXIA FONOLÓGICA":
                fon_prob = 70.0
                mix_prob = 31.5
            elif test_diag == "DISLEXIA SUPERFICIAL (VISUAL)":
                sup_prob = 70.0
                mix_prob = 31.5
            elif test_diag == "DISLEXIA MIXTA":
                fon_prob = 65.0
                sup_prob = 65.0
                mix_prob = 70.0
        
        max_prob = max(fon_prob, sup_prob, mix_prob)
        gravedad = "NORMAL"
        if max_prob >= 70:
            gravedad = "SEVERO"
        elif max_prob >= 40:
            gravedad = "MODERADO"
        elif max_prob >= 15:
            gravedad = "LEVE"

        narrativo = "El paciente no tiene sesiones de juego registradas todavía."
        if test_diag:
            narrativo = f"El paciente no tiene sesiones de juego, pero cuenta con un Test Clínico PROLEC-R finalizado con diagnóstico de **{test_diag}** (Riesgo: {gravedad}). Se sugiere iniciar la terapia sugerida."

        # Analisis de grafemas de errores del test dislexia
        confusiones_conteos = {
            'B_V': 0, 'G_J': 0, 'C_S_Z': 0, 'MIRROR': 0, 'H_OMISSION': 0, 'R_RR': 0, 'OTHER': 0
        }
        for err in dis_test_errors:
            word = err.get("word", "")
            read_as = err.get("read_as", "")
            conf_type = analyze_confusion(word, read_as)
            confusiones_conteos[conf_type] += 1
            
        total_errores = sum(confusiones_conteos.values())
        analisis_errores = []
        for conf_key, cantidad in confusiones_conteos.items():
            if cantidad > 0:
                pct = round((cantidad / total_errores * 100), 1)
                analisis_errores.append({
                    "clave": conf_key,
                    "label": ETIQUETAS[conf_key],
                    "cantidad": cantidad,
                    "porcentaje": pct
                })
        analisis_errores.sort(key=lambda x: x["cantidad"], reverse=True)
        error_principal_key = analisis_errores[0]["clave"] if analisis_errores else 'OTHER'
        template_exercises = EXERCISE_TEMPLATES.get(error_principal_key, EXERCISE_TEMPLATES['OTHER'])
        
        # Recomendaciones basadas en errores del test
        test_failed_words = [(e.get("word", ""), e.get("read_as", "")) for e in dis_test_errors if e.get("state") == "incorrect" and e.get("read_as") != "no se escuchó la palabra" and e.get("word")]
        custom_exercises = [generate_custom_exercise(w, r, error_principal_key) for w, r in test_failed_words[:3]]
        
        ejercicios_recomendados = []
        ejercicios_recomendados.extend(custom_exercises)
        for templ_ex in template_exercises:
            if len(ejercicios_recomendados) >= 5:
                break
            if not any(ex["palabra"].lower() == templ_ex["palabra"].lower() for ex in ejercicios_recomendados):
                ejercicios_recomendados.append(templ_ex)
        while len(ejercicios_recomendados) < 5 and template_exercises:
            ejercicios_recomendados.append(template_exercises[0])

        return {
            "paciente": paciente_info,
            "nicknames": [],
            "latest_test": latest_test,
            "diagnostico": {
                "fonologica_prob": round(fon_prob, 1),
                "superficial_prob": round(sup_prob, 1),
                "mixta_prob": round(mix_prob, 1),
                "nivel_gravedad": gravedad,
                "narrativo": narrativo
            },
            "metricas": {"totales": 0, "correctas": 0, "incorrectas": 0, "precision": 0, "tiempo_medio": 0},
            "analisis_errores": analisis_errores,
            "ejercicios_recomendados": ejercicios_recomendados,
            "historico_juegos": []
        }
        
    with connection.cursor() as cursor:
        # Obtener todos los intentos individuales detallados
        format_strings = ','.join(['%s'] * len(player_ids))
        cursor.execute(f"""
            SELECT ga.word_shown, ga.answer_given, ga.is_correct, ga.reaction_time_ms, ga.error_type, gs.game_type, gs.level, ga.fecha_registro
            FROM game_attempts ga
            JOIN game_sessions gs ON ga.session_id = gs.id
            WHERE gs.player_id IN ({format_strings})
            ORDER BY ga.fecha_registro DESC
        """, player_ids)
        attempts_rows = cursor.fetchall()
        
        # Obtener el historico consolidado de resultados de juego (para tabla general)
        cursor.execute("""
            SELECT nombre_juego, respuestas_correctas, respuestas_incorrectas, tiempo_jugado_segundos, estrellas_ganadas, niveles_completados, fecha_resultado
            FROM resultadosjuegos
            WHERE id_paciente = %s
            ORDER BY fecha_resultado DESC
        """, [patient_id])
        resultados_rows = cursor.fetchall()

    # Formatear historico consolidado
    historico_juegos = []
    for r in resultados_rows:
        total = (r[1] or 0) + (r[2] or 0)
        prec = round((r[1] or 0) / total * 100, 1) if total > 0 else 0
        historico_juegos.append({
            "juego": r[0],
            "correctas": r[1] or 0,
            "incorrectas": r[2] or 0,
            "tiempo": r[3] or 0,
            "estrellas": r[4] or 0,
            "niveles_completados": r[5] or 0,
            "precision": prec,
            "fecha": r[6].strftime("%d/%m/%Y %H:%M") if r[6] else "N/A"
        })

    # Procesar intentos detallados
    total_intentos = len(attempts_rows)
    total_correctas = sum(1 for row in attempts_rows if row[2] == 1)
    total_incorrectas = total_intentos - total_correctas
    precision_total = round((total_correctas / total_intentos * 100), 1) if total_intentos > 0 else 0
    tiempos = [row[3] for row in attempts_rows if row[3] > 0]
    tiempo_promedio = round(sum(tiempos) / len(tiempos) / 1000, 2) if tiempos else 0 # en segundos
    
    # Agrupar metricas por rutas lectoras clinicas
    phon_attempts = [row for row in attempts_rows if row[5] in PHONOLOGICAL_GAMES]
    sup_attempts = [row for row in attempts_rows if row[5] in SUPERFICIAL_GAMES]
    mixed_attempts = [row for row in attempts_rows if row[5] in MIXED_GAMES]
    
    phon_total = len(phon_attempts)
    phon_inc = sum(1 for row in phon_attempts if row[2] == 0)
    phon_error_rate = phon_inc / phon_total if phon_total > 0 else 0
    phon_avg_rt = sum(row[3] for row in phon_attempts) / phon_total if phon_total > 0 else 0
    
    sup_total = len(sup_attempts)
    sup_inc = sum(1 for row in sup_attempts if row[2] == 0)
    sup_error_rate = sup_inc / sup_total if sup_total > 0 else 0
    sup_avg_rt = sum(row[3] for row in sup_attempts) / sup_total if sup_total > 0 else 0
    
    # Algoritmo de inferencia del Sistema Experto para Probabilidades de Dislexia
    rt_phon_factor = min(20.0, max(0.0, (phon_avg_rt - 3000.0) / 150.0)) if phon_total > 0 else 0
    rt_sup_factor = min(20.0, max(0.0, (sup_avg_rt - 3000.0) / 150.0)) if sup_total > 0 else 0
    
    fon_prob = 0
    if phon_total > 0:
        base_prob = (phon_error_rate / 0.5) * 80.0
        fon_prob = min(100.0, max(0.0, base_prob + rt_phon_factor))
        
    sup_prob = 0
    if sup_total > 0:
        base_prob = (sup_error_rate / 0.5) * 80.0
        sup_prob = min(100.0, max(0.0, base_prob + rt_sup_factor))
        
    mix_prob = 0
    if phon_total > 0 and sup_total > 0:
        mix_prob = min(fon_prob, sup_prob) * 1.1
        mix_prob = min(100.0, mix_prob)
    elif phon_total > 0 or sup_total > 0:
        mix_prob = (fon_prob + sup_prob) * 0.4
        
    # Integrar diagnóstico clínico del Test de Dislexia (PROLEC-R) en las probabilidades
    test_diag = latest_test["diagnostico"] if latest_test else None
    if test_diag:
        if test_diag == "DISLEXIA FONOLÓGICA":
            fon_prob = max(70.0, min(100.0, fon_prob + 30.0))
            mix_prob = (fon_prob + sup_prob) * 0.45
        elif test_diag == "DISLEXIA SUPERFICIAL (VISUAL)":
            sup_prob = max(70.0, min(100.0, sup_prob + 30.0))
            mix_prob = (fon_prob + sup_prob) * 0.45
        elif test_diag == "DISLEXIA MIXTA":
            fon_prob = max(65.0, min(100.0, fon_prob + 25.0))
            sup_prob = max(65.0, min(100.0, sup_prob + 25.0))
            mix_prob = max(70.0, min(100.0, mix_prob + 35.0))
        elif test_diag == "LECTOR NORMAL":
            # Si el test clínico formal descarta patología, moderamos el indicador del juego
            fon_prob *= 0.5
            sup_prob *= 0.5
            mix_prob *= 0.5
        
    # Clasificar gravedad
    max_prob = max(fon_prob, sup_prob, mix_prob)
    gravedad = "NORMAL"
    if max_prob >= 70:
        gravedad = "SEVERO"
    elif max_prob >= 40:
        gravedad = "MODERADO"
    elif max_prob >= 15:
        gravedad = "LEVE"
        
    # Analisis especifico de confusiones de grafemas
    incorrect_attempts = [row for row in attempts_rows if row[2] == 0]
    confusiones_conteos = {
        'B_V': 0, 'G_J': 0, 'C_S_Z': 0, 'MIRROR': 0, 'H_OMISSION': 0, 'R_RR': 0, 'OTHER': 0
    }
    
    for row in incorrect_attempts:
        conf_type = analyze_confusion(row[0], row[1], row[4])
        confusiones_conteos[conf_type] += 1
        
    # Incorporar errores detectados en el test clínico de dislexia
    for err in dis_test_errors:
        word = err.get("word", "")
        read_as = err.get("read_as", "")
        conf_type = analyze_confusion(word, read_as)
        confusiones_conteos[conf_type] += 1
        
    total_errores = sum(confusiones_conteos.values())
    analisis_errores = []
    
    for conf_key, cantidad in confusiones_conteos.items():
        if cantidad > 0:
            pct = round((cantidad / total_errores * 100), 1)
            analisis_errores.append({
                "clave": conf_key,
                "label": ETIQUETAS[conf_key],
                "cantidad": cantidad,
                "porcentaje": pct
            })
            
    analisis_errores.sort(key=lambda x: x["cantidad"], reverse=True)
    error_principal_key = analisis_errores[0]["clave"] if analisis_errores else 'OTHER'
    template_exercises = EXERCISE_TEMPLATES.get(error_principal_key, EXERCISE_TEMPLATES['OTHER'])
    
    # Extraer palabras reales o pseudopalabras falladas en el Test PROLEC-R para recomendación
    test_failed_words = []
    for err in dis_test_errors:
        word = err.get("word", "")
        state = err.get("state", "")
        read_as = err.get("read_as", "")
        # Ignorar omisiones simples
        if state == "incorrect" and read_as != "no se escuchó la palabra" and word:
            test_failed_words.append((word, read_as))
            
    custom_exercises = []
    # Generar hasta 3 ejercicios correctivos personalizados
    for word, read_as in test_failed_words[:3]:
        ex = generate_custom_exercise(word, read_as, error_principal_key)
        custom_exercises.append(ex)
        
    # Rellenar hasta 5 ejercicios usando la plantilla correspondiente a su fallo principal
    ejercicios_recomendados = []
    ejercicios_recomendados.extend(custom_exercises)
    for templ_ex in template_exercises:
        if len(ejercicios_recomendados) >= 5:
            break
        # Evitar duplicar palabras
        if not any(ex["palabra"].lower() == templ_ex["palabra"].lower() for ex in ejercicios_recomendados):
            ejercicios_recomendados.append(templ_ex)
            
    # Si aun faltan ejercicios para llegar a 5, completar con plantilla
    while len(ejercicios_recomendados) < 5 and template_exercises:
        ejercicios_recomendados.append(template_exercises[0])
        
    if analisis_errores:
        letra_falla = analisis_errores[0]["label"]
        pct_falla = analisis_errores[0]["porcentaje"]
    else:
        letra_falla = "Ninguna"
        pct_falla = 0

    # Construir narrativo
    if total_intentos == 0:
        narrativo = "El paciente aun no ha realizado sesiones de juego suficientes para elaborar un diagnostico clinico."
    else:
        narrativo = f"El Sistema Experto analizo {total_intentos} intentos de juego con una precision global del {precision_total}%.\n\n"
        if gravedad == "NORMAL":
            narrativo += "Desempeño normalizado. El paciente no muestra indicadores de riesgo clinicamente significativos de dislexia evolutiva en las rutas fonologica o visual."
        else:
            narrativo += f"Se detecta un nivel de riesgo **{gravedad}** de dislexia. "
            if mix_prob > 60:
                narrativo += "El perfil clínico indica un patron de **Dislexia Mixta**. Presenta dificultades marcadas tanto en la ruta de decodificacion silabica (auditiva) como en la ruta visual-ortografica (directa)."
            elif fon_prob > sup_prob:
                narrativo += "El perfil clinico indica una tendencia hacia **Dislexia Fonologica**. El paciente muestra mayor dificultad en la correspondencia grafema-fonema, el analisis de silabas y la audicion reflexiva de las palabras."
            else:
                narrativo += "El perfil clinico indica una tendencia hacia **Dislexia Superficial (Visual)**. El paciente presenta problemas con la memoria visual de las palabras, cometiendo fallos ortograficos recurrentes y confusiones de grafias con fonemas similares."
                
            if test_diag:
                narrativo += f"\n\nEste diagnóstico fue enriquecido e integrado con los resultados clínicos de su Test PROLEC-R realizado el {latest_test['fecha']}, el cual arrojó un diagnóstico formal de **{test_diag}**."
                
            if pct_falla > 0:
                narrativo += f"\n\nEl analisis de grafemas destaca que la mayor dificultad se concentra en **{letra_falla}** (representando el {pct_falla}% del total de sus errores). Se sugiere reconfigurar los niveles para centrar la practica del paciente en estas grafias."

    return {
        "paciente": paciente_info,
        "nicknames": nicknames,
        "diagnostico": {
            "fonologica_prob": round(fon_prob, 1),
            "superficial_prob": round(sup_prob, 1),
            "mixta_prob": round(mix_prob, 1),
            "nivel_gravedad": gravedad,
            "narrativo": narrativo
        },
        "metricas": {
            "totales": total_intentos,
            "correctas": total_correctas,
            "incorrectas": total_incorrectas,
            "precision": precision_total,
            "tiempo_medio": tiempo_promedio
        },
        "analisis_errores": analisis_errores,
        "ejercicios_recomendados": ejercicios_recomendados,
        "historico_juegos": historico_juegos
    }


def generate_styled_excel(patient_id):
    """
    Genera un archivo Excel (.xlsx) altamente estilizado con openpyxl utilizando la paleta
    de colores y guia de estilos de NeuroGym.
    """
    profile = get_patient_profile(patient_id)
    if not profile:
        return None
        
    wb = Workbook()
    
    # 1. ESTILOS DE DISEÑO
    font_family = "Segoe UI"
    
    # Tipografias
    font_title = Font(name=font_family, size=16, bold=True, color="1E1B4B")
    font_subtitle = Font(name=font_family, size=11, italic=True, color="555555")
    font_section = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=10, color="000000")
    font_bold_data = Font(name=font_family, size=10, bold=True, color="000000")
    font_risk = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    # Rellenos (Fills)
    fill_navy = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    fill_teal = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    fill_cyan = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_risk_severo = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    fill_risk_moderado = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    fill_risk_leve = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    fill_risk_normal = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    # Bordes y Alineaciones
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # --- HOJA 1: RESUMEN CLINICO ---
    ws1 = wb.active
    ws1.title = "Resumen Clinico"
    ws1.views.sheetView[0].showGridLines = True
    
    # Titulo de la Hoja
    ws1["A1"] = "NeuroGym — Reporte del Sistema Experto IA"
    ws1["A1"].font = font_title
    ws1["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')} para fines de diagnostico y terapia."
    ws1["A2"].font = font_subtitle
    
    # Seccion 1: Ficha del Paciente
    ws1.merge_cells("A4:D4")
    ws1["A4"] = "DATOS DEL PACIENTE"
    ws1["A4"].font = font_section
    ws1["A4"].fill = fill_navy
    ws1["A4"].alignment = align_left
    
    ws1["A5"] = "Nombre Completo:"
    ws1["B5"] = profile["paciente"]["nombre"]
    ws1["C5"] = "Edad Actual:"
    ws1["D5"] = f"{profile['paciente']['edad']} años"
    
    ws1["A6"] = "Institucion/Colegio:"
    ws1["B6"] = profile["paciente"]["colegio"]
    ws1["C6"] = "Fecha Registro:"
    ws1["D6"] = profile["paciente"]["fecha_registro"]
    
    # Seccion 2: Perfil de Riesgo
    ws1.merge_cells("A8:D8")
    ws1["A8"] = "DIAGNOSTICO DE RIESGO DE DISLEXIA"
    ws1["A8"].font = font_section
    ws1["A8"].fill = fill_navy
    ws1["A8"].alignment = align_left
    
    ws1["A9"] = "Tipo de Evaluacion"
    ws1["B9"] = "Probabilidad"
    ws1["C9"] = "Nivel de Gravedad"
    ws1["D9"] = "Estado Clinico"
    
    for col in ["A", "B", "C", "D"]:
        ws1[f"{col}9"].font = font_header
        ws1[f"{col}9"].fill = fill_teal
        ws1[f"{col}9"].alignment = align_center
        ws1[f"{col}9"].border = thin_border
        
    diag = profile["diagnostico"]
    gravedad = diag["nivel_gravedad"]
    fill_g = fill_risk_normal
    if gravedad == "SEVERO":
        fill_g = fill_risk_severo
    elif gravedad == "MODERADO":
        fill_g = fill_risk_moderado
    elif gravedad == "LEVE":
        fill_g = fill_risk_leve
        
    ws1["A10"] = "Dislexia Fonologica"
    ws1["B10"] = f"{diag['fonologica_prob']}%"
    ws1["C10"] = gravedad if diag['fonologica_prob'] > 20 else "NORMAL"
    ws1["D10"] = "Requiere intervencion" if diag['fonologica_prob'] > 40 else "Rango controlado"
    
    ws1["A11"] = "Dislexia Superficial (Visual)"
    ws1["B11"] = f"{diag['superficial_prob']}%"
    ws1["C11"] = gravedad if diag['superficial_prob'] > 20 else "NORMAL"
    ws1["D11"] = "Requiere intervencion" if diag['superficial_prob'] > 40 else "Rango controlado"
    
    ws1["A12"] = "Dislexia Mixta"
    ws1["B12"] = f"{diag['mixta_prob']}%"
    ws1["C12"] = gravedad if diag['mixta_prob'] > 20 else "NORMAL"
    ws1["D12"] = "Atencion inmediata" if diag['mixta_prob'] > 40 else "Rango controlado"
    
    for r in range(10, 13):
        for c in ["A", "B", "C", "D"]:
            cell = ws1[f"{c}{r}"]
            cell.font = font_data
            cell.border = thin_border
            if c in ["B", "C"]:
                cell.alignment = align_center
            if c == "C" and cell.value != "NORMAL":
                cell.fill = fill_g
                cell.font = font_risk
                
    # Seccion 3: Metricas Clave
    ws1.merge_cells("A14:D14")
    ws1["A14"] = "INDICADORES CLAVE DE DESEMPEÑO"
    ws1["A14"].font = font_section
    ws1["A14"].fill = fill_navy
    ws1["A14"].alignment = align_left
    
    m = profile["metricas"]
    ws1["A15"] = "Intentos Totales:"
    ws1["B15"] = m["totales"]
    ws1["C15"] = "Precision Promedio:"
    ws1["D15"] = f"{m['precision']}%"
    
    ws1["A16"] = "Respuestas Correctas:"
    ws1["B16"] = m["correctas"]
    ws1["C16"] = "Respuestas Incorrectas:"
    ws1["D16"] = m["incorrectas"]
    
    ws1["A17"] = "Tiempo Medio por Intento:"
    ws1["B17"] = f"{m['tiempo_medio']} s"
    ws1["C17"] = "Nivel Maximo Alcanzado:"
    ws1["D17"] = "Nivel 10" if m["totales"] > 0 else "N/A"
    
    # Seccion 4: Evaluación Clínica (Test PROLEC-R)
    ws1.merge_cells("A19:D19")
    ws1["A19"] = "EVALUACIÓN CLÍNICA (TEST PROLEC-R)"
    ws1["A19"].font = font_section
    ws1["A19"].fill = fill_navy
    ws1["A19"].alignment = align_left
    
    latest_test = profile.get("latest_test")
    if latest_test:
        ws1["A20"] = "Fecha del Test:"
        ws1["B20"] = latest_test["fecha"]
        ws1["C20"] = "Diagnóstico Clínico:"
        ws1["D20"] = latest_test["diagnostico"]
        
        ws1["A21"] = "Palabras Aciertos:"
        ws1["B21"] = f"{latest_test['a_p']} / 40"
        ws1["C21"] = "Eficiencia Palabras (IL-P):"
        ws1["D21"] = latest_test["il_p"]
        
        ws1["A22"] = "Pseudopalabras Aciertos:"
        ws1["B22"] = f"{latest_test['a_ps']} / 40"
        ws1["C22"] = "Eficiencia Pseudos (IL-PS):"
        ws1["D22"] = latest_test["il_ps"]
    else:
        ws1.merge_cells("A20:D22")
        ws1["A20"] = "No se registran evaluaciones del Test de Dislexia (PROLEC-R) para este paciente."
        ws1["A20"].alignment = align_center
        ws1["A20"].font = font_subtitle
        
    # Seccion 5: Informe Narrativo
    ws1.merge_cells("A24:D24")
    ws1["A24"] = "INFORME NARRATIVO DEL SISTEMA EXPERTO"
    ws1["A24"].font = font_section
    ws1["A24"].fill = fill_navy
    ws1["A24"].alignment = align_left
    
    ws1.merge_cells("A25:D30")
    ws1["A25"] = diag["narrativo"]
    ws1["A25"].font = font_data
    ws1["A25"].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    for r in [5, 6, 15, 16, 17, 20, 21, 22]:
        for c in ["A", "B", "C", "D"]:
            cell = ws1[f"{c}{r}"]
            cell.border = thin_border
            if not latest_test and r in [20, 21, 22]:
                continue
            if c in ["A", "C"]:
                cell.font = font_bold_data
            else:
                cell.font = font_data
                
    # --- HOJA 2: HISTORIAL DE SESIONES ---
    ws2 = wb.create_sheet(title="Historial de Sesiones")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "NeuroGym — Historial de Resultados Consolidados"
    ws2["A1"].font = font_title
    
    headers2 = ["Juego", "Respuestas Correctas", "Respuestas Incorrectas", "Precision (%)", "Tiempo Jugado (s)", "Estrellas", "Niveles Completados", "Fecha Fin"]
    for col_idx, text in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx, s in enumerate(profile["historico_juegos"], 4):
        row_data = [
            s["juego"], s["correctas"], s["incorrectas"], 
            f"{s['precision']}%", s["tiempo"], s["estrellas"], 
            s["niveles_completados"], s["fecha"]
        ]
        fill_row = fill_zebra if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            if fill_row.fill_type:
                cell.fill = fill_row
            if c_idx > 1:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # --- HOJA 3: DETALLE DE ERRORES ---
    ws3 = wb.create_sheet(title="Detalle de Errores")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3["A1"] = "NeuroGym — Analisis de Grafemas e Intentos Fallidos"
    ws3["A1"].font = font_title
    
    ws3["A3"] = "Frecuencia de Errores por Letra / Silaba"
    ws3["A3"].font = Font(name=font_family, size=11, bold=True, color="1E1B4B")
    
    headers3_1 = ["Patron de Fallo", "Cantidad", "Porcentaje sobre Errores"]
    for col_idx, text in enumerate(headers3_1, 1):
        cell = ws3.cell(row=4, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = thin_border
        
    next_row = 5
    for r_idx, err in enumerate(profile["analisis_errores"], 5):
        row_data = [err["label"], err["cantidad"], f"{err['porcentaje']}%"]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            if c_idx > 1:
                cell.alignment = align_center
        next_row += 1
        
    next_row += 1
    ws3.cell(row=next_row, column=1, value="Listado Historico de Intentos con Errores").font = Font(name=font_family, size=11, bold=True, color="1E1B4B")
    next_row += 1
    
    headers3_2 = ["Juego / Categoria", "Nivel", "Palabra Mostrada", "Respuesta del Paciente", "Clasificacion IA", "Fecha Registro"]
    for col_idx, text in enumerate(headers3_2, 1):
        cell = ws3.cell(row=next_row, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    # Obtener los intentos detallados de la base de datos
    with connection.cursor() as cursor:
        cursor.execute("SELECT id FROM game_players WHERE id_paciente = %s", [patient_id])
        player_ids = [row[0] for row in cursor.fetchall()]
        
        if player_ids:
            format_strings = ','.join(['%s'] * len(player_ids))
            cursor.execute(f"""
                SELECT ga.word_shown, ga.answer_given, ga.is_correct, ga.reaction_time_ms, ga.error_type, gs.game_type, gs.level, ga.fecha_registro
                FROM game_attempts ga
                JOIN game_sessions gs ON ga.session_id = gs.id
                WHERE gs.player_id IN ({format_strings})
                ORDER BY ga.fecha_registro DESC
            """, player_ids)
            attempts_rows = cursor.fetchall()
        else:
            attempts_rows = []
            
    start_errs_row = next_row + 1
    err_index = start_errs_row
    
    for row in attempts_rows:
        if row[2] == 0: # Solo incorrectos
            conf = analyze_confusion(row[0], row[1], row[4])
            conf_label = ETIQUETAS.get(conf, "Otro error")
            
            row_data = [
                str(row[5]).capitalize(), row[6], row[0], row[1], conf_label,
                row[7].strftime("%d/%m/%Y %H:%M") if row[7] else "N/A"
            ]
            fill_row = fill_zebra if err_index % 2 == 0 else PatternFill(fill_type=None)
            
            for c_idx, val in enumerate(row_data, 1):
                cell = ws3.cell(row=err_index, column=c_idx, value=val)
                cell.font = font_data
                cell.border = thin_border
                if fill_row.fill_type:
                    cell.fill = fill_row
                if c_idx in [2, 3, 4, 6]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
            err_index += 1

    # --- HOJA 4: RECOMENDACIONES Y EJERCICIOS ---
    ws4 = wb.create_sheet(title="Ejercicios Correctivos")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4["A1"] = "NeuroGym — Ejercicios Sugeridos para Tratamiento de Grafemas"
    ws4["A1"].font = font_title
    ws4["A2"] = "Basado en los patrones de fallos ortograficos recurrentes del paciente, la IA sugiere entrenar con estas palabras:"
    ws4["A2"].font = font_subtitle
    
    headers4 = ["Palabra Sugerida", "Opciones del Juego (Distractores)", "Juego Recomendado"]
    for col_idx, text in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx, ex in enumerate(profile["ejercicios_recomendados"], 5):
        row_data = [
            ex["palabra"], ", ".join(ex["opciones"]), ex["juego_recomendado"]
        ]
        fill_row = fill_zebra if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for c_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            if fill_row.fill_type:
                cell.fill = fill_row
            if c_idx == 2:
                cell.alignment = align_center
                cell.font = font_bold_data
            else:
                cell.alignment = align_left

    # AUTO-AJUSTAR ANCHO DE LAS COLUMNAS EN TODAS LAS HOJAS
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.coordinate in ["A1", "A2", "A4", "A8", "A14", "A19", "A20", "A24", "A25"] or len(val_str) > 50:
                    continue
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Guardar en buffer en memoria o escribir temporalmente
    temp_dir = os.path.normpath(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'neurogym', 'backend', 'media', 'reportes'))
    os.makedirs(temp_dir, exist_ok=True)
    
    file_name = f"Reporte_ExpertoIA_Paciente_{patient_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    wb.save(file_path)
    
    return file_path, file_name
